#!/usr/bin/python
#-*- coding:utf-8 -*-
#@Author：zhuxiujie

# 用例可配置
from openpyxl import load_workbook
from tool.read_config import ReadConfig

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        mode = ReadConfig().read_config('case.config','MODE','mode')
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2,sheet.max_row+1):
            sub_data = {}
            sub_data['case_id'] = sheet.cell(i,1).value
            sub_data['method'] = sheet.cell(i,2).value
            sub_data['url'] = sheet.cell(i,3).value
            sub_data['data'] = sheet.cell(i, 4).value
            sub_data['expected'] = sheet.cell(i, 5).value
            test_data.append(sub_data)
        if mode == 'all':
            final_data = test_data
        else:
            final_data = []
            for item in test_data:
                if item['case_id'] in eval(mode):
                    final_data.append(item)
        return final_data
print(DoExcel('python.xlsx','python').get_data())